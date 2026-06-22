"""
excel_export.py — Snap2Sheet
Builds a styled 3-sheet Excel workbook.
"""

import logging
from typing import Any, Dict, List
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)

# Colours
BG_DARK   = "0A0C10"
BG_HEAD   = "1C2130"
BG_SUB    = "252D40"
BG_HW     = "0D3B4D"
BG_PR     = "0D3D22"
BG_ALT    = "141820"
C_CYAN    = "00E5FF"
C_GREEN   = "7BFFB0"
C_WHITE   = "E8EAF2"
C_MUTED   = "7A8099"
C_DIM     = "434A63"
C_BORDER  = "2A3045"


def build_excel(fields: List[Dict[str, Any]], out_path: str) -> str:
    wb = Workbook()
    _sheet_formdata(wb, fields)
    _sheet_summary(wb, fields)
    _sheet_readme(wb)
    wb.save(out_path)
    logger.info(f"Excel saved: {out_path} ({len(fields)} fields)")
    return out_path


def _border(color=C_BORDER):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(color): return PatternFill("solid", fgColor=color)
def _font(color=C_WHITE, bold=False, size=10):
    return Font(name="Arial", color=color, bold=bold, size=size)
def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _sheet_formdata(wb: Workbook, fields):
    ws = wb.active
    ws.title = "Form Data"
    ws.sheet_properties.tabColor = C_CYAN
    ws.freeze_panes = "A4"

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 52
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 20

    # Title
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "Snap2Sheet — Extracted Form Data"
    c.font, c.fill, c.alignment = _font(C_CYAN, True, 15), _fill(BG_HEAD), _align("center")
    ws.row_dimensions[1].height = 30

    # Meta
    ws.merge_cells("A2:E2")
    c = ws["A2"]
    hw = sum(1 for f in fields if f.get("type") == "handwritten")
    pr = len(fields) - hw
    c.value = f"Generated: {datetime.now():%Y-%m-%d %H:%M}  |  Total: {len(fields)}  |  Handwritten: {hw}  |  Printed: {pr}"
    c.font, c.fill, c.alignment = _font(C_MUTED, False, 9), _fill(BG_HEAD), _align("center")
    ws.row_dimensions[2].height = 16

    # Header row
    headers = ["#", "Field Name", "Extracted Value", "Type", "Timestamp"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font, c.fill, c.alignment, c.border = _font(C_WHITE, True, 10), _fill(BG_SUB), _align("center"), _border()
    ws.row_dimensions[3].height = 20

    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    for ri, f in enumerate(fields, 1):
        row  = ri + 3
        hw   = f.get("type") == "handwritten"
        bg   = BG_HW if hw else (BG_ALT if ri % 2 == 0 else BG_DARK)
        vals = [ri, f.get("key",""), f.get("value",""),
                "✍️ Handwritten" if hw else "🖨️ Printed", ts]
        colors = [C_MUTED, C_CYAN if hw else C_GREEN, C_WHITE, C_CYAN if hw else C_GREEN, C_MUTED]
        aligns = ["center","left","left","center","center"]
        for ci, (v, col, al) in enumerate(zip(vals, colors, aligns), 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.font, c.fill, c.alignment, c.border = (
                _font(col, ci==2, 10), _fill(bg), _align(al, wrap=ci==3), _border()
            )
        ws.row_dimensions[row].height = 20

    ws.auto_filter.ref = f"A3:E{3+len(fields)}"


def _sheet_summary(wb: Workbook, fields):
    ws = wb.create_sheet("Summary")
    ws.sheet_properties.tabColor = C_GREEN
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 58

    ws.merge_cells("A1:B1")
    c = ws["A1"]
    c.value = "Form Summary — Field → Value"
    c.font, c.fill, c.alignment = _font(C_CYAN, True, 13), _fill(BG_HEAD), _align("left")
    ws.row_dimensions[1].height = 26

    for ci, lbl in enumerate(["Field Name", "Value"], 1):
        c = ws.cell(row=2, column=ci, value=lbl)
        c.font, c.fill, c.alignment, c.border = _font(C_WHITE, True, 10), _fill(BG_SUB), _align("center"), _border()
    ws.row_dimensions[2].height = 18

    for ri, f in enumerate(fields, 1):
        row = ri + 2
        hw  = f.get("type") == "handwritten"
        bg  = BG_HW if hw else (BG_ALT if ri%2==0 else BG_DARK)
        ac  = C_CYAN if hw else C_GREEN

        k = ws.cell(row=row, column=1, value=f.get("key",""))
        k.font, k.fill, k.alignment, k.border = _font(ac, True, 10), _fill(bg), _align(), _border()

        v = ws.cell(row=row, column=2, value=f.get("value",""))
        v.font, v.fill, v.alignment, v.border = _font(C_WHITE, False, 10), _fill(bg), _align(wrap=True), _border()
        ws.row_dimensions[row].height = 20


def _sheet_readme(wb: Workbook):
    ws = wb.create_sheet("README")
    ws.sheet_properties.tabColor = "FF6B35"
    ws.column_dimensions["A"].width = 80
    lines = [
        ("Snap2Sheet — Workbook Guide", True, 14, C_CYAN),
        ("", False, 9, C_DIM),
        ("Sheet: Form Data", True, 11, C_CYAN),
        ("Every extracted field as a row: Field Name | Value | Type | Timestamp", False, 10, C_WHITE),
        ("Auto-filter is enabled — click column headers to sort/filter.", False, 10, C_MUTED),
        ("", False, 9, C_DIM),
        ("Sheet: Summary", True, 11, C_GREEN),
        ("Compact two-column view: Field Name → Value. Best for copy-pasting.", False, 10, C_WHITE),
        ("", False, 9, C_DIM),
        ("Colour coding:", True, 10, C_WHITE),
        ("  Cyan rows  = Handwritten fields (LayoutLMv3 / Tesseract HW)", False, 10, C_CYAN),
        ("  Green rows = Printed fields (Tesseract OCR)", False, 10, C_GREEN),
        ("", False, 9, C_DIM),
        ("Engines:", True, 10, C_WHITE),
        ("  LayoutLMv3 (nielsr/layoutlmv3-finetuned-funsd) — spatial form understanding", False, 10, C_MUTED),
        ("  Tesseract OCR — raw text extraction + fallback parsing", False, 10, C_MUTED),
        ("", False, 9, C_DIM),
        (f"Generated: {datetime.now():%Y-%m-%d %H:%M:%S}", False, 9, C_DIM),
    ]
    for ri, (txt, bold, size, col) in enumerate(lines, 1):
        c = ws.cell(row=ri, column=1, value=txt)
        c.font, c.fill, c.alignment = _font(col, bold, size), _fill(BG_HEAD), _align(v="center")
        ws.row_dimensions[ri].height = 18 if txt else 7
